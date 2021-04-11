import openpyxl
from openpyxl.styles import PatternFill  # 导入填充模块
from openpyxl.comments import Comment  # 导入批注模块

# file = input("请输入excel文件路径名：")         #定位到现有的excel
# wb1 = openpyxl.load_workbook(file)

# 定位到现有的excel
wb1 = openpyxl.load_workbook("E:\python\\working.xlsx")
bo1 = wb1["work2"]  # 定位到要使用的表

str1 = "requirement"
str2 = "comment"
str3 = "shall"
colum1 = 5  # 需要用到的列    列E
colum2 = 6  # 需要用到的列    列F
colum3 = 9  # 需要用到的列    列I
colum4 = 10  # 需要用到的列    列J
colum5 = 12  # 该列存放shall的个数     列L

fille = PatternFill('solid', fgColor='FFBB00')  # 设置填充颜色为 橙色
fille1 = PatternFill('solid', fgColor='FFFFFF')  # 设置填充颜色为 白色

num = 1

while 1:  # 计算有效行数
    cell = bo1.cell(row=num, column=1).value
    if cell:
        num = num + 1
    else:
        print(f'总行数：{num - 1}')
        break

# 开始扫描时将所有单元格颜色设置成白色
for row in range(1, bo1.max_row):
    for clmn in range(1, bo1.max_column):
        bo1.cell(row, clmn).fill = fille1

for a in range(2, num):  # 2到num行
    i = 0  # i用来计算shall的个数
    stra = bo1.cell(a, colum2).value  # 取出第F列的值
    if stra == str1:  # 如果是requirement
        # 1.首先计算shall的个数
        stra1 = bo1.cell(a, colum1).value  # 取出第E列的值
        if stra1:
            stra1 = stra1.split(' ')
            for value in stra1:
                if value == str3:  # 如果是shall则i的值加1
                    i += 1
            # bo1.cell(a, colum5).value = i             #将shall个数显示到第L列
            # 若shall个数不是1则填充颜色表示有误,加上背景色并加上批注
            if i != 1:
                location = 'E' + str(a)  # 记录单元格位置
                print(location)
                bo1[location].fill = fille
                comment = Comment('Object Type is requirement but the number of "shall" is not one!', 'zhangfan')
                comment.width = 300
                comment.height = 50
                bo1[location].comment = comment
        else:
            location = 'E' + str(a)
            print(location)
            bo1[location].fill = fille
            comment = Comment('Modified HLR must be not empty!', 'zhangfan')
            comment.width = 300
            comment.height = 50
            bo1[location].comment = comment
        # 2.第I列值为非空，并且值只能是Yes或No
        stra1 = bo1.cell(a, colum3).value
        cell = bo1.cell(a, colum4).value
        if (stra1 == 'Yes'):  # 若J列在I列为yes情况下为空，则填充颜色，下条if语句判断字符串为非空，反方向代替，该步只是为下步else做铺垫
            if cell:
                x = 1
            else:
                location = 'J' + str(a)
                print(location)
                bo1[location].fill = fille
                comment = Comment('Derived is Yes so you Rationale must be not empty!', 'zhangfan')
                comment.width = 300
                comment.height = 50
                bo1[location].comment = comment
        elif (stra1 == 'No'):
            if cell:  # 若J列在I列为No情况下不为空，则填充颜色
                location = 'J' + str(a)
                print(location)
                bo1[location].fill = fille
                comment = Comment('Derived is No so you Rationale must be empty!', 'zhangfan')
                comment.width = 300
                comment.height = 50
                bo1[location].comment = comment

        else:  # I列为空或者不是Yes或No的情况,表示错误，填充颜色
            location = 'I' + str(a)
            print(location)
            bo1[location].fill = fille
            comment = Comment('Derived is empty or its content is not in Yes or No!', 'zhangfan')
            comment.width = 300
            comment.height = 50
            bo1[location].comment = comment
            if cell:  # 第J列不为空则错误，填充颜色
                location = 'J' + str(a)
                print(location)
                bo1[location].fill = fille
                comment = Comment('Derived is error,so Rationale must be empty!', 'zhangfan')
                comment.width = 300
                comment.height = 50
                bo1[location].comment = comment
    elif stra == str2:  # 如果是comment
        # 1.首先计算是否有shall，有几个
        stra1 = bo1.cell(a, colum1).value  # 取出第E列的值
        if stra1:  # 不为空
            stra1 = stra1.split(' ')
            for value in stra1:
                if (value == str3):  # 如果是shall则i的值加1
                    i += 1
            # comment不能有shall，有的话填充颜色
            if i:
                location = 'E' + str(a)
                print(location)
                bo1[location].fill = fille
                comment = Comment('Object Type is comment so the number of "shall" must be zero!', 'zhangfan')
                comment.width = 300
                comment.height = 50
                bo1[location].comment = comment
        else:
            location = 'E' + str(a)
            print(location)
            bo1[location].fill = fille
            comment = Comment('Modified HLR must be not empty!', 'zhangfan')
            comment.width = 300
            comment.height = 50
            bo1[location].comment = comment
        # 2.第I列值为空,J列值为空
        cell = bo1.cell(a, colum3).value
        if cell:  # 在第F列为comment情况下第I列不为空则填充颜色
            location = 'I' + str(a)
            print(location)
            bo1[location].fill = fille
            comment = Comment('Object Type is comment so Derived must be empty!', 'zhangfan')
            comment.width = 300
            comment.height = 50
            bo1[location].comment = comment
        cell = bo1.cell(a, colum4).value
        if cell:  # 在第F列为comment情况下第J列不为空则填充颜色
            location = 'J' + str(a)
            print(location)
            bo1[location].fill = fille
            comment = Comment('Object Type is comment so Rationale must be empty!', 'zhangfan')
            comment.width = 300
            comment.height = 50
            bo1[location].comment = comment
    else:  # 第F列不是requirement也不是comment，填充整行元素
        for j in range(1, bo1.max_column):
            bo1.cell(row=a, column=j).fill = fille
        location = 'F' + str(a)
        print(location)
        bo1[location].fill = fille
        comment = Comment('Object Type is error or empty!', 'zhangfan')
        comment.width = 300
        comment.height = 50
        bo1[location].comment = comment

wb1.save("E:\python\\working.xlsx")
input("输入任意字符结束")
